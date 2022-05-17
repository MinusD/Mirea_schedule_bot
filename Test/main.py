import datetime

import requests
import openpyxl
import pickle
import re

import matplotlib.pyplot as plt
import numpy as np
from PIL import Image, ImageDraw, ImageFont
from urllib.request import urlopen

from bs4 import BeautifulSoup
import src.cfgs.system_config as scfg


def get_wind_slug(speed: float):
    s = 'Ураган'
    if speed <= 0.2:
        s = 'Штиль'
    elif speed <= 1.5:
        s = 'Тихий'
    elif speed <= 3.3:
        s = 'Лёгкий'
    elif speed <= 5.4:
        s = 'Слабый'
    elif speed <= 7.9:
        s = 'Умеренный'
    elif speed <= 10.7:
        s = 'Свежий'
    elif speed <= 13.8:
        s = 'Сильный'
    elif speed <= 17.1:
        s = 'Крепкий'
    elif speed <= 20.7:
        s = 'Очень крепкий'
    elif speed <= 24.4:
        s = 'Шторм'
    elif speed <= 28.4:
        s = 'Сильный шторм'
    elif speed <= 32.6:
        s = 'Жестокий шторм'
    return s


def get_wind_deg_slug(deg: int):
    s = 'северный'
    if deg <= 22.5:
        s = 'северный'
    elif deg <= 67.5:
        s = 'северо-восточный'
    elif deg <= 112.5:
        s = 'восточный'
    elif deg <= 157.5:
        s = 'юго-восточный'
    elif deg <= 202.5:
        s = 'южный'
    elif deg <= 247.5:
        s = 'юго-западный'
    elif deg <= 292.5:
        s = 'западный'
    elif deg <= 292.5:
        s = 'северо-западный'
    return s


if __name__ == '__main__':
    image = Image.open('../src/assets/weather_pattern.jpg')
    # img = image.resize((400, 500))
    img = image.resize((400, 260))
    img = img.convert('RGB')
    idraw = ImageDraw.Draw(img)
    title = ImageFont.truetype(font='../src/assets/lato.ttf', size=30)
    font = ImageFont.truetype(font='../src/assets/lato.ttf', size=18)
    # title2 = ImageFont.truetype(size=30)
    idraw.text((10, 10), 'Погода в Москве на 5 дней', font=title, fill="white")

    appid = '483841295963b30a56e7679ae38f99e1'
    response = requests.get("http://api.openweathermap.org/data/2.5/weather",
                            params={'q': 'moscow,ru', 'units': 'metric', 'APPID': appid, 'lang': 'ru'})
    # response = requests.get(
    #     'https://api.openweathermap.org/data/2.5/weather?q=&appid=483841295963b30a56e7679ae38f99e1&units=metric'
    #     '&lang=ru')
    info = response.json()
    print(info)
    pattern = '{}, температура: {}°C\nДавление: {} мм рт. сб., влажность: {}%\nВетер: {}, {} м/с, {}'
    img_name = info['weather'][0]['icon']
    # weather_image = requests.get(, stream=True).content
    im = Image.open(urlopen('https://openweathermap.org/img/wn/{}@4x.png'.format(img_name)))
    img.paste(im, (95, 20), im.convert('RGBA'))
    status = info['weather'][0]['description'].capitalize()
    temp = str(int(info['main']['temp_min'])) + ' - ' + str(int(info['main']['temp_max']))
    pressure = str(int(float(info['main']['pressure']) / 1.33))
    humidity = str(info['main']['humidity'])
    wind_speed = info['wind']['speed']
    wind_slug = get_wind_slug(float(wind_speed)).lower()
    wind_deg_slug = get_wind_deg_slug(info['wind']['speed'])
    weather = pattern.format(status, temp, pressure, humidity, wind_slug, wind_speed, wind_deg_slug)
    idraw.text((10, 185), weather, font=font, fill="white")

    img.save('../data/weather_card.png')

    # temp = info["main"]["temp"]
    # print(temp)

    # now = datetime.datetime.now()
    # print(now.isocalendar())

    # page = requests.get(scfg.CORONA_STAT_URL + '/country/russia')  # Получаем страницу
    # soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
    # result = soup.find('table', {'class': 'table table-bordered small'}).findAll('tr')
    # # days = active = cured = died = cases = stats = []
    # days = []
    # active = []
    # cured = []
    # died = []
    # cases = []
    # stats = []
    # ml = 1000000
    # for i in range(1, 11):
    #     days.append(result[i].find('th').getText())
    #     for a in result[i].findAll('td'):
    #         stats.append(int(a.getText().split(' ')[1]))
    # for i in range(0, len(stats), 4):
    #     active.append(stats[i] / ml)
    # for i in range(1, len(stats), 4):
    #     cured.append(stats[i] / ml)
    # for i in range(2, len(stats), 4):
    #     died.append(stats[i] / ml)
    # for i in range(3, len(stats), 4):
    #     cases.append(stats[i] / ml)
    # var = days, active, cured, died, cases
    # population_by_continent = {
    #     'Активных': active,
    #     'Вылечено': cured,
    #     'Умерло': died,
    #     # 'europe': cases,
    # }
    # for i in range(len(days)):
    #     days[i] = days[i][:-5]
    #     # if i % 2:
    #     #     days[i] = '_'
    # # print(days)
    # days = list(reversed(days))
    # fig, ax = plt.subplots()
    # ax.stackplot(days, population_by_continent.values(),
    #              labels=population_by_continent.keys(), alpha=0.8)
    #
    # ax.legend(loc='upper left')
    # ax.set_title('World population')
    # ax.set_xlabel('Year')
    # ax.set_ylabel('Number of people (millions)')
    #
    # plt.show()

    # page = requests.get(scfg.CORONA_STAT_URL + '/country/russia')  # Получаем страницу
    # soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
    # result = soup.find('table', {'class': 'table table-bordered small'}).findAll('tr')
    # active = []
    # cured = []
    # died = []
    # cases = []
    # stats = []
    # for i in range(1, 11):
    #     for a in result[i].findAll('td'):
    #         stats.append(int(a.getText().split(' ')[1]))
    # for i in range(0, len(stats), 4):
    #     active.append(stats[i])
    # for i in range(1, len(stats), 4):
    #     cured.append(stats[i])
    # for i in range(2, len(stats), 4):
    #     died.append(stats[i])
    # for i in range(3, len(stats), 4):
    #     cases.append(stats[i])
    # print(active)
    # print(cured)
    # print(died)
    # print(cases)

    # print(result)

    # region = 'Москва'
    # page = requests.get(scfg.CORONA_STAT_URL + '/country/russia')  # Получаем страницу
    # soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
    # result = soup.findAll('div', {'class': 'c_search_row'})
    # d = ''
    # for x in result:
    #     tmp = x.find('span', 'small').find('a')
    #     if region.title() in tmp.getText().split(' '):
    #         rg = tmp.getText()
    #         d = tmp.get('href')
    # print('URL:', scfg.CORONA_STAT_URL + d)
    # page = requests.get(scfg.CORONA_STAT_URL + d)  # Получаем страницу
    # soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
    # result = soup.find(string='Прогноз заражения на 10 дней').find_parent('div',
    #                                                                       {'class': 'border rounded mt-3 mb-3 p-3'})
    # status = result.find('h6', 'text-muted').getText()[:-17]
    #
    # print(status)
    # print(rg)
    # data = result.findAll('div', {'class': 'col col-6 col-md-3 pt-4'})
    # plus = []
    # value = []
    # for i in range(4):
    #     value.append(data[i].find('div', 'h2').getText())
    #     plus.append(data[i].find('span', {'class': 'font-weight-bold'}).getText())
    # print(value)
    # print(plus)
    # result = soup.find(string="Институт информационных технологий").find_parent("div").find_parent("div").findAll(
    #     'a', {'class': 'uk-link-toggle'})

    # course_pattern = r'([1-3]) курс'
    # for x in result:
    #     course = x.find('div', 'uk-link-heading').text.lower().strip()
    #     course_number = re.match(course_pattern, course)
    #     if course_number:
    #         course_number = course_number.group(1)
    #         f = open(f'{scfg.DATA_DIR}{scfg.SCHEDULE_BASE_NAME}{course_number}.xlsx', "wb")
    #         link = x.get('href')
    #         resp = requests.get(link)  # запрос по ссылке
    #         f.write(resp.content)  # Записываем контент в файл
    #         f.close()
    # self.last_schedule_file_update = time.time()
    # Debug('Update schedule files', key='UPD')
    # self._parse_schedule_file()

    # a = [i for i in range(10)]
    # for i in range(len(a)):
    #     if i % 2 == 0:
    #         del a[i]
    # for i in range(len(a)):
    #     print(a[i])
    # s = '123\n321'
    # a = '123'
    # d = s.split('\n')
    # d.append('213')
    # print(' - '.join(d))
    # print(' - '.join(a.split('\n')))

    # data = ['Структуры и алгоритмы обработки данных',
    #         '4,8,12,16 н. Физика (1 п/г)\n4,8,12,16 н. Физика (2 п/г)',
    #         'кр. 12 н. Объектно-ориентированное программирование',
    #         'кр. 1,3 н. Ознакомительная практика']
    #
    # custom_week_pattern_1 = r'([\d\,]+) н. ([^\\]+)'
    # custom_week_pattern = r'кр. ([\d\,]+) н. ([^\\]+)'
    #
    # for a in data:
    #     s = re.search(custom_week_pattern, a)
    #     if s:
    #         d = s.group(1).split(',')
    #         print(d)
    #         print(s.group(2))
    #     else:
    #         s = re.search(custom_week_pattern_1, a)
    #         if s:
    #             d = s.group(1).split(',')
    #             print('есть', d)
    #             print(s.group(2))

    # print(re.search(custom_week_pattern, text1))
    # print(re.search(custom_week_pattern, text2))
    # print(re.search(custom_week_pattern, text3))
    # print(re.search(custom_week_pattern, text4))

    # group_slug = "ИКБО-30-21"
    # book = openpyxl.load_workbook(
    #     f'../{scfg.DATA_DIR}{scfg.SCHEDULE_BASE_NAME}{1}.xlsx')  # открытие файла
    # sheet = book.active  # активный лист
    # num_cols = sheet.max_column  # количество столбцов
    # num_rows = sheet.max_row  # количество строк
    # data = []
    # last_group = 0
    # for i in range(6, num_cols):
    #     if last_group >= 4:
    #         last_group = -1
    #         continue
    #     tmp = []
    #     for j in range(2, 75):
    #         v = str(sheet.cell(column=i, row=j).value)
    #         if j == 2 and re.match(scfg.GROUP_PATTERN, v):
    #             last_group = 0
    #         tmp.append(v)
    #     if last_group != -1:
    #         data.append(tmp)
    #         last_group += 1
    # # print(data)
    # for a in data:
    #     # print(a)
    #     print(a[0])
    # with open('data.pickle', 'wb') as f:
    #     pickle.dump(data, f, fix_imports=True)
    # del data

    # print('dd')
    #
    # a = int(input())
    # cell = sheet.cell(row=2, column=(i + 1))
    #
    # if sheet.cell(row=2, column=(i + 1)).value == group_slug:
    #     print(i + 1)
    pass

# page = requests.get("https://www.mirea.ru/schedule/")
#
# soup = BeautifulSoup(page.text, "html.parser")
# # print(soup)
# # print(soup.find('div'))
# # result = soup.find(string="Институт информационных технологий").find_parent("div").find_parent("div").findAll('a')
# result = soup.find(string="Институт информационных технологий").find_parent("div").find_parent("div").findAll('a', {
#     'class': 'uk-link-toggle'})
# # print(result)
# for x in result:
#     if x.find('div', 'uk-link-heading').text.strip() == '1 курс':
#         f = open("file.xlsx", "wb")
#         link = x.get('href')
#         resp = requests.get(link)  # запрос по ссылке
#         f.write(resp.content)

# book = openpyxl.load_workbook("file.xlsx")  # открытие файла
# sheet = book.active  # активный лист
# num_cols = sheet.max_column  # количество столбцов
# num_rows = sheet.max_row  # количество строк
# for i in range(num_cols):
#     group_name_pattern = r''
#     cell = sheet.cell(row=2, column=i+1).value
#     print(cell)
#         # cell = sheet.cell(row=1, column=1).value
#         # print(cell)
#     # print(x.find('div', 'uk-link-heading').text.strip())
#     # if x.find(string='1 Курс'):
#     #     print(x)
#     # if x.find('a', 'uk-link-toggle')
#     # if ...:  # среди всех ссылок найти нужную
#     # f = open("file.xlsx", "wb")  # открываем файл для записи, в режиме wb
#     # resp = requests.get(...)  # запрос по ссылке
#     # f.write(resp.content)
