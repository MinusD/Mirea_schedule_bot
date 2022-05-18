# Встроенные библиотеки
import re
import time
from typing import Any

import matplotlib.pyplot as plt
import openpyxl
import requests

# Сторонние библиотеки
import vk_api
from bs4 import BeautifulSoup
from vk_api import VkUpload
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id

# Конфиги
import src.cfgs.main_config as cfg
import src.cfgs.system_config as scfg

# Самописные модули
from src.helper_module import *
from src.sql_database import *


class VkBot:
    def __init__(self) -> None:
        """
        Конструктор
        """
        self.vk_session = vk_api.VkApi(token=scfg.TOKEN)
        self.vk = self.vk_session.get_api()
        self.longpoll = VkLongPoll(self.vk_session)
        self.users_to_set_group: set = set()
        self.users_to_set_teacher: set = set()
        self.users_to_get_teacher: list = []
        self.last_schedule_file_update: time
        self.schedule_data: list
        if scfg.UPDATE_SCHEDULE_FILE_ON_START:
            self._update_schedule_file()
        else:
            self._parse_schedule_file()
        Debug('Bot init', key='SRT')

    def start_listen(self) -> None:
        """
        Слушатель событий вк

        :return:
        """
        Debug('Start listen', key='SRT')
        for event in self.longpoll.listen():
            if event.type == VkEventType.MESSAGE_NEW and event.text and event.to_me:
                Debug('from {} text = \'{}\''.format(event.user_id, event.text), key='MSG')
                self._command_handler(event.user_id, event.text.lower())

    def _command_handler(self, user_id: int, text: str) -> None:
        """
        Обработчик команд

        :param user_id:
        :param text:
        :return:
        """
        for i in range(len(self.users_to_get_teacher)):
            if self.users_to_get_teacher[i][0] == user_id:
                match text:
                    case cfg.BTN_SCHEDULE_TODAY:
                        self._show_today_teacher_schedule(user_id, self.users_to_get_teacher[i][1])
                    case cfg.BTN_SCHEDULE_TOMORROW:
                        self._show_today_teacher_schedule(user_id, self.users_to_get_teacher[i][1], 1)
                    case cfg.BTN_SCHEDULE_WEEK:
                        self._show_teacher_week_schedule(user_id, self.users_to_get_teacher[i][1])
                    case cfg.BTN_SCHEDULE_NEXT_WEEK:
                        self._show_teacher_week_schedule(user_id, self.users_to_get_teacher[i][1], 1)
                del self.users_to_get_teacher[i]
                return

        match text:
            case cfg.CMD_START:
                user_data = self.vk.users.get(user_id=user_id)[0]
                self._send_message(user_id, cfg.ABOUT_TEXT.format(user_data['first_name']))
                Debug(f'Start new user: {user_data["first_name"]} {user_data["last_name"]}')
                self._add_user_to_edit_group_list(user_id)
                return
            case cfg.CMD_SCHEDULE:
                self._show_schedule_keyboard(user_id)
                return
            case cfg.BTN_SCHEDULE_TODAY:
                self._show_today_schedule(user_id)
                return
            case cfg.BTN_SCHEDULE_TOMORROW:
                self._show_tomorrow_schedule(user_id)
                return
            case cfg.BTN_SCHEDULE_WEEK:
                self._show_week_schedule(user_id)
                return
            case cfg.BTN_SCHEDULE_NEXT_WEEK:
                self._show_week_schedule(user_id, week_delta=1)
                return
            case cfg.BTN_WHAT_WEEK:
                self._show_current_week(user_id)
                return
            case cfg.BTN_WHAT_GROUP:
                self._show_user_group(user_id)
                return
            case cfg.BTN_HELP:
                self._show_help_message(user_id)
                return
            case cfg.CMD_CORONA:
                self._show_corona_all_stat(user_id)
                return

        combo_cmd = text.split(' ')
        match combo_cmd[0]:
            case cfg.CMD_SCHEDULE:
                if len(combo_cmd) == 2:
                    if combo_cmd[1].lower() in cfg.WEEK_DAYS_INFINITIVE_SLUGS:
                        self._show_week_day_schedule(user_id, combo_cmd[1].lower())
                    else:
                        self._edit_user_group(user_id, combo_cmd[1])
                    return
                elif len(combo_cmd) == 3:
                    if combo_cmd[1].lower() in cfg.WEEK_DAYS_INFINITIVE_SLUGS:
                        if self._validate_group_slug(combo_cmd[2]):
                            self._edit_user_group(user_id, combo_cmd[2])
                            self._show_week_day_schedule(user_id, combo_cmd[1].lower())
                        else:
                            self._send_message(user_id, cfg.INVALID_GROUP_TEXT)
                        return
            case cfg.CMD_FIND_TEACHER:
                self._show_teacher_keyboard(user_id, combo_cmd[1:])
                return
            case cfg.CMD_CORONA:
                self._show_corona_local_data(user_id, combo_cmd[1:])
                return

        if str(user_id) in self.users_to_set_group:
            self._edit_user_group(user_id, text)
            return

        if str(user_id) in self.users_to_set_teacher:
            self._show_teacher_period_keyboard(user_id, text)
            return

        Debug(f'Command {text} not found for id: {user_id}', key='CNF')
        self._send_message(user_id, cfg.INVALID_COMMAND_TEXT.format(cfg.BTN_HELP.title()))

    def _get_week_schedule(self, group: str, date: datetime.datetime, with_reformat: bool = True) -> list:
        """
        Возвращает расписанию на неделю, дату которой передали

        :param group:
        :param date:
        :param with_reformat:
        :return:
        """
        now = date.isocalendar()
        week = now.week + scfg.WEEK_DELTA
        week_even = (week + 1) % 2  # Является ли неделя чётной
        column = self._get_group_column(group)
        out = []
        tmp = []
        for i in range(2 + week_even, len(self.schedule_data[column]), 2):  # Каждый второй, со смещением по недели
            tmp.append(
                [self.schedule_data[column][i],  # Предмет
                 self.schedule_data[column + 1][i],  # Вид занятий
                 self.schedule_data[column + 2][i],  # Преподаватель
                 self.schedule_data[column + 3][i]]  # Кабинет
            )
            if len(tmp) == 6:
                out.append(tmp)
                tmp = []

        for i in range(len(out)):
            for j in range(6):
                out[i][j][0] = self._reformat_subject_name(out[i][j][0], week_number=week,
                                                           ignore_weeks=(not with_reformat))
                out[i][j][1] = self._reformat_double_pair(out[i][j][1])
                out[i][j][2] = self._reformat_double_pair(out[i][j][2])
                out[i][j][3] = self._reformat_double_pair(out[i][j][3])
        Debug(f'Getting {group} week schedule', key='GET')
        return out

    def _get_day_schedule(self, group: str, date: datetime.datetime) -> list:
        """
        Возвращает массив с расписанием на текущий день

        :param group: Номер группы
        :param date: День
        :return:
        """
        week = self._get_week_schedule(group, date)
        week_index = date.isocalendar().weekday - 1
        if week_index == 6:
            return [[] * 4] * 6
        Debug(f'Getting {group} day schedule', key='GET')
        return week[week_index]

    def _get_user_group(self, user_id: int) -> str or None:
        """
        Получает группу пользователя или ошибка

        :param user_id:
        :return: Номер группы или None
        """
        group = Database().fetch_one(table=scfg.TABLE_NAME, condition=f'user_id = {user_id}')
        if group:
            Debug(f'Find {user_id} group: {group}', key='FND')
            return group[1]
        else:
            self._send_message(user_id, cfg.CURRENT_GROUP_ERROR_TEXT.format(cfg.CMD_SCHEDULE.title()))
        return None

    def _get_current_week(self) -> int:
        """
        Возвращает номер текущей недели

        :return: учебная неделя
        """
        Debug(f'Getting current week', key='GET')
        return datetime.datetime.now().isocalendar().week + scfg.WEEK_DELTA

    def _get_group_column(self, group) -> int or None:
        """
        Ищет столбец группы в расписании

        :param group:
        :return: Индекс столбца или None
        """
        for i in range(0, len(self.schedule_data), 4):
            if self.schedule_data[i][0] == group:
                Debug(f'Find in file group: {group} column: {i + 1}', key='FND')
                return i
        return None

    def _get_string_date(self, date: datetime.datetime, with_week_day: bool = False) -> str:
        """
        Преобразует дату в строку с датой

        :param date:
        :param with_week_day:
        :return:
        """
        result = ''
        if with_week_day:
            result += cfg.WEEK_DAYS_SLUGS[date.isocalendar().weekday - 1] + " "
        result += str(date.day) + " " + cfg.MONTHS_SLUGS[date.month % 12 - 1]
        return result

    def _get_teacher_full_name(self, teacher: str) -> set[str]:
        """
        Получение полного имени преподавателей из расписания

        :param teacher: Фамилия преподавателя
        :return: Множество полных имён с заданной фамилией
        """
        result = set()
        for i in range(2, len(self.schedule_data), 4):
            for j in range(2, len(self.schedule_data[i])):
                tmp = self.schedule_data[i][j].split('\n')
                if len(tmp) > 0:
                    if tmp[0].split(' ')[0] == teacher:
                        result.add(tmp[0] if tmp[0][-1] == '.' else tmp[0] + '.')  # Исправление косяков расписания
                    elif tmp[-1].split(' ')[0] == teacher:

                        result.add(tmp[-1] if tmp[-1][-1] == '.' else tmp[-1] + '.')
        Debug(f'Getting {teacher} fullname', key='GET')
        return result

    def _get_teacher_week_schedule(self, teacher: str, date: datetime.datetime, with_reformat: bool = True) -> list:
        """
        Возвращает расписание преподавателя на указанную неделю

        :param teacher:
        :param date:
        :param with_reformat:
        :return:
        """
        now = date.isocalendar()
        week = now.week + scfg.WEEK_DELTA
        week_even = (week + 1) % 2  # Является ли неделя чётной
        out = []
        tmp = []

        for j in range(2 + week_even, len(self.schedule_data[0]), 2):
            para = []  # одна пара
            for i in range(2, len(self.schedule_data), 4):  # Слева на права
                tmp_teachers = self.schedule_data[i][j].split('\n')  # для сдвоенных пар
                if len(tmp_teachers) > 0:
                    t1 = tmp_teachers[0] if tmp_teachers[0][-1] == '.' else tmp_teachers[0] + '.'
                    t2 = tmp_teachers[-1] if tmp_teachers[-1][-1] == '.' else tmp_teachers[-1] + '.'
                    if t1 == teacher:
                        para = [
                            self.schedule_data[i - 2][j].split('\n')[0],  # Предмет
                            self.schedule_data[i - 1][j].split('\n')[0],  # Вид
                            self.schedule_data[i - 2][0],  # Группа
                            self.schedule_data[i + 1][j].split('\n')[0]  # Аудитория
                        ]
                        break
                    elif t2 == teacher:
                        para = [
                            self.schedule_data[i - 2][j].split('\n')[-1],  # Предмет
                            self.schedule_data[i - 1][j].split('\n')[-1],  # Вид
                            self.schedule_data[i - 2][0],  # Группа
                            self.schedule_data[i + 1][j].split('\n')[-1]  # Аудитория
                        ]
                        break
                    # Останавливаем смешение вправо, если нашли
            tmp.append(para)  # Добавляем пару, даже если она пустая
            if (j - week_even) % 12 == 0:
                out.append(tmp)
                tmp = []

        for i in range(len(out)):
            for j in range(6):
                if len(out[i][j]) > 1:
                    out[i][j][0] = self._reformat_subject_name(out[i][j][0], week_number=week,
                                                               ignore_weeks=(not with_reformat))
        return out

    def _get_day_teacher_schedule(self, teacher: str, date: datetime.datetime) -> list:
        """
        Возвращает расписание преподавателя на переданный день

        :param teacher:
        :param date:
        :return:
        """
        week = self._get_teacher_week_schedule(teacher, date)
        week_index = date.isocalendar().weekday - 1
        if week_index == 6:
            return [[] * 4] * 6
        return week[week_index]

    def _get_corona_stat(self, extra_url: str = '') -> tuple[str, list[Any], list[Any]]:
        """
        Возвращает статистику коронавируса на сегодня определённой области

        :param extra_url:
        :return:
        """
        page = requests.get(scfg.CORONA_STAT_URL + extra_url)  # Получаем страницу
        soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
        result = soup.find(string='Прогноз заражения на 10 дней').find_parent('div', {
            'class': 'border rounded mt-3 mb-3 p-3'})
        status = result.find('h6', 'text-muted').getText()[:-17]
        data = result.findAll('div', {'class': 'col col-6 col-md-3 pt-4'})
        plus = [] * 4
        value = [] * 4
        for i in range(4):
            value.append(data[i].find('div', 'h2').getText())
            plus.append(data[i].find('span', {'class': 'font-weight-bold'}).getText())
        return status, value, plus

    def _get_corona_all_stat(self) -> tuple[list[Any], list[float], list[float], list[float], list[float]]:
        """
        Возвращает статистику коронавируса на последние 10 дней

        :return:
        """
        page = requests.get(scfg.CORONA_STAT_URL + '/country/russia')  # Получаем страницу
        soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
        result = soup.find('table', {'class': 'table table-bordered small'}).findAll('tr')
        days = []
        active = []
        cured = []
        died = []
        cases = []
        stats = []
        ml = 1000000

        for i in range(1, 11):
            days.append(result[i].find('th').getText())
            for a in result[i].findAll('td'):
                stats.append(int(a.getText().split(' ')[1]))
        for i in range(0, len(stats), 4):
            active.append(stats[i] / ml)
        for i in range(1, len(stats), 4):
            cured.append(stats[i] / ml)
        for i in range(2, len(stats), 4):
            died.append(stats[i] / ml)
        for i in range(3, len(stats), 4):
            cases.append(stats[i] / ml)

        days = list(reversed(days))
        active = list(reversed(active))
        cured = list(reversed(cured))
        died = list(reversed(died))
        cases = list(reversed(cases))

        return days, active, cured, died, cases

    def _show_corona_all_stat(self, user_id: int) -> None:
        """
        Показывает статистику коронавируса на сегодня и выводит график

        :param user_id:
        :return:
        """
        days, active, cured, died, cases = self._get_corona_all_stat()
        graf_data = {
            'Активных': active,
            'Вылечено': cured,
            'Умерло': died,
        }
        for i in range(len(days)):
            days[i] = days[i][:-5]
        fig, ax = plt.subplots()
        ax.stackplot(days, graf_data.values(),
                     labels=graf_data.keys(), alpha=0.8)
        ax.legend(loc='upper left')
        ax.set_title(cfg.GRAF_HEADER_TEXT)
        ax.set_ylabel(cfg.GRAF_LABEL_TEXT)
        fig.savefig(scfg.DATA_DIR + scfg.GRAF_FILENAME)
        upload = VkUpload(self.vk_session)
        attachments = []
        photo = upload.photo_messages(scfg.DATA_DIR + scfg.GRAF_FILENAME)[0]
        attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
        self._send_message_with_attachments(user_id=user_id,
                                            text=self._reformat_corona_data('Россия', self._get_corona_stat()),
                                            attachments=attachments)

    def _show_corona_local_data(self, user_id: int, region_list: list) -> None:
        """
        Показывает статистику коронавируса по региону

        :param user_id:
        :param region_list:
        :return:
        """
        if len(region_list) > 0:
            region = region_list[0].title()
            page = requests.get(scfg.CORONA_STAT_URL + '/country/russia')  # Получаем страницу
            soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
            result = soup.findAll('div', {'class': 'c_search_row'})
            d = ''
            rg = cfg.CORONA_REGION_DEFAULT
            for x in result:
                tmp = x.find('span', 'small').find('a')
                if region.title() in tmp.getText().split(' '):
                    rg = tmp.getText()
                    d = tmp.get('href')
                    break
            self._send_message(user_id, self._reformat_corona_data(rg, self._get_corona_stat(d)))

    def _show_today_teacher_schedule(self, user_id: int, teacher: str, day_delta: int = 0) -> None:
        """
        Выводит расписания на сегодня день(со смещением)

        :param user_id:
        :param teacher:
        :param day_delta: Сколько дней вперёд
        :return:
        """
        now = datetime.datetime.now() + datetime.timedelta(days=day_delta)
        teacher = self._reformat_teacher_name(teacher)
        if self._validate_teacher_name(teacher):
            schedule = self._get_day_teacher_schedule(teacher, now)
            self._send_message(user_id=user_id, text=self._reformat_day_schedule(schedule, now, teacher_header=teacher))

    def _show_teacher_week_schedule(self, user_id: int, teacher: str, week_delta: int = 0) -> None:
        """
        Выводит расписание преподавателя на неделю

        :param user_id:
        :param teacher:
        :param week_delta:
        :return:
        """
        now = datetime.datetime.now() + datetime.timedelta(weeks=week_delta)
        day_date = now - datetime.timedelta(days=now.isocalendar().weekday - 1)
        result = ''
        teacher = self._reformat_teacher_name(teacher)
        if self._validate_teacher_name(teacher):
            schedule = self._get_teacher_week_schedule(teacher, now)
            for i in range(6):
                result += self._reformat_day_schedule(schedule[i], date=day_date, teacher_header=teacher,
                                                      week_format=True)
                day_date += datetime.timedelta(days=1)
            self._send_message(user_id, result)

    def _show_teacher_period_keyboard(self, user_id: int, teacher: str) -> None:
        """
        Показывает клавиатуру для выбора периода расписания преподавателя

        :param user_id:
        :param teacher: Полное имя преподавателя
        :return:
        """
        tmp = teacher.split(' ')
        if len(tmp) == 2:
            teacher = tmp[0].title() + ' ' + tmp[1].upper()
            if self._validate_teacher_name(teacher):
                Debug(f'Show period for {teacher}', key='SHW')
                # Создаём клавиатуру
                keyboard = VkKeyboard(one_time=True)
                keyboard.add_button(cfg.BTN_SCHEDULE_TODAY, color=VkKeyboardColor.POSITIVE)
                keyboard.add_button(cfg.BTN_SCHEDULE_TOMORROW, color=VkKeyboardColor.NEGATIVE)
                keyboard.add_line()
                keyboard.add_button(cfg.BTN_SCHEDULE_WEEK, color=VkKeyboardColor.PRIMARY)
                keyboard.add_button(cfg.BTN_SCHEDULE_NEXT_WEEK, color=VkKeyboardColor.PRIMARY)

                self._clear_wait_lists(user_id)
                self._add_user_to_get_teacher_list(user_id, teacher)
                self._send_message(user_id, text=cfg.TEACHER_SELECT_PERIOD_TEXT.format(teacher),
                                   custom_keyboard=keyboard)
                return
        self._send_message(user_id, text=cfg.TEACHER_SELECT_ERROR_TEXT)

    def _show_teacher_keyboard(self, user_id: int, teacher: list = None):
        """
        Показывает клавиатуру выбора преподавателя, либо выбора периода, для показа расписания

        :param user_id:
        :param teacher:
        :return:
        """
        name = ''
        if len(teacher) == 2:
            name = teacher[0].title() + ' ' + teacher[1].upper()
        elif len(teacher) == 1:  # Только фамилия
            tmp = []
            for a in self._get_teacher_full_name(teacher[0].title()):
                tmp.append(a)
            if len(tmp) == 1:  # Если 1, то
                name = tmp[0]
            elif len(tmp) > 1:
                self._add_user_to_set_teacher_list(user_id)  # Добавляем пользователя в список ожидания
                keyboard = VkKeyboard(one_time=True)
                for i in range(len(tmp)):
                    keyboard.add_button(tmp[i], color=VkKeyboardColor.SECONDARY)
                    if i % 2 and i != len(tmp) - 1:  # Каждый второй, но не последний
                        keyboard.add_line()
                self._send_message(user_id=user_id, text=cfg.TEACHER_SELECT_TEXT, custom_keyboard=keyboard)
                Debug(f'Show teacher keyboard id {user_id}')
                return
        if len(name) > 1:
            self._show_teacher_period_keyboard(user_id, name)
            return
        self._send_message(user_id, cfg.TEACHER_SELECT_ERROR_TEXT)

    def _show_help_message(self, user_id: int) -> None:
        """
        Отправляет подсказку с командами

        :param user_id:
        :return:
        """
        self._send_message(user_id, cfg.HELP_TEXT)

    def _show_week_day_schedule(self, user_id: int, day: str) -> None:
        """
        Показывает расписание на определённый день недели

        :param user_id:
        :param day:
        :return:
        """
        group = self._get_user_group(user_id)
        if group:
            date = datetime.datetime.now()
            if self._get_current_week() % 2 == 0:  # Если она не чётная
                date -= datetime.timedelta(weeks=1)
            odd = self._get_week_schedule(group=group, date=date, with_reformat=False)
            date += datetime.timedelta(weeks=1)
            even = self._get_week_schedule(group=group, date=date, with_reformat=False)
            index = cfg.WEEK_DAYS_INFINITIVE_SLUGS.index(day)
            o = self._reformat_day_schedule(data=odd[index], with_header=False)  # Нечётный день
            e = self._reformat_day_schedule(data=even[index], with_header=False)  # Чётный день
            result = cfg.ODD_DAY_PATTERN.format(day.title()) + o + '\n\n' + cfg.EVEN_DAY_PATTERN.format(day.title()) + e
            self._send_message(user_id, result)

    def _show_today_schedule(self, user_id: int) -> None:
        """
        Выводит расписание на сегодняшний день

        :param user_id:
        :return:
        """
        now = datetime.datetime.now()
        group = self._get_user_group(user_id)
        if group:
            schedule = self._get_day_schedule(group, now)
            self._send_message(user_id=user_id, text=self._reformat_day_schedule(schedule, now))

    def _show_tomorrow_schedule(self, user_id: int) -> None:
        """
        Выводит расписание на завтрашний день

        :param user_id:
        :return:
        """
        now = datetime.datetime.now() + datetime.timedelta(days=1)
        group = self._get_user_group(user_id)
        if group:
            schedule = self._get_day_schedule(group, now)
            Debug(f'Show {group} tomorrow schedule for id: {user_id}', key='SHW')
            self._send_message(user_id=user_id, text=self._reformat_day_schedule(schedule, now))

    def _show_week_schedule(self, user_id: int, week_delta: int = 0) -> None:
        """
        Выводит расписание на текущую неделю

        :param user_id:
        :return:
        """
        now = datetime.datetime.now() + datetime.timedelta(weeks=week_delta)
        day_date = now - datetime.timedelta(days=now.isocalendar().weekday - 1)
        group = self._get_user_group(user_id)
        result = ''
        if group:
            schedule = self._get_week_schedule(group, now)
            for i in range(6):
                result += self._reformat_day_schedule(schedule[i], date=day_date, week_format=True)
                day_date += datetime.timedelta(days=1)
            Debug(f'Show {group} week schedule for id: {user_id}', key='SHW')
            self._send_message(user_id, result)

    def _show_current_week(self, user_id: int) -> None:
        """
        Выводит пользователю номер текущей недели

        :param user_id:
        :return:
        """
        self._send_message(user_id, cfg.CURRENT_WEEK_TEXT.format(self._get_current_week()))

    def _show_user_group(self, user_id: int) -> None:
        """
        Выводит пользователю номер выбранной группы или ошибку

        :param user_id:
        :return:
        """
        group = self._get_user_group(user_id)
        if group:
            self._send_message(user_id, cfg.CURRENT_GROUP_TEXT.format(group))

    def _show_schedule_keyboard(self, user_id: int) -> None:
        """
        Показать клавиатуру выбора расписания

        :param user_id:
        :return:
        """
        Debug(f'Show schedule keyboard for id: {user_id}')
        self._send_message(user_id, cfg.SCHEDULE_SELECT_TEXT, keyboard=1)

    def _add_user_to_edit_group_list(self, user_id) -> None:
        """
        Добавляет пользователя в список обновления группы

        :param user_id:
        :return:
        """
        self.users_to_set_group.add(str(user_id))
        Debug(f'User add to set group list, uid: {user_id}', key='SET')

    def _add_user_to_set_teacher_list(self, user_id) -> None:
        """
        Добавляет пользователя в список выбора преподавателя

        :param user_id:
        :return:
        """
        self.users_to_set_teacher.add(str(user_id))
        Debug(f'User add to set teacher list, uid: {user_id}', key='SET')

    def _add_user_to_get_teacher_list(self, user_id, teacher: str) -> None:
        """
        Добавляет пользователя в список выбора преподавателя

        :param user_id:
        :return:
        """
        self.users_to_get_teacher.append([user_id, teacher])
        Debug(f'User add to get teacher list, uid: {user_id}', key='SET')

    def _edit_user_group(self, user_id: int, group_slug: str) -> None:
        """
        Изменить группу пользователя или выдать ошибку

        :param user_id:
        :param group_slug:
        :return:
        """
        group_slug = group_slug.upper()
        if self._validate_group_slug(group_slug):
            db = Database()
            if not db.fetch_one(table=scfg.TABLE_NAME, condition=f'user_id = {user_id}'):
                db.insert_one(table=scfg.TABLE_NAME, data=[user_id, group_slug])
                Debug(f'Set user group: {group_slug} uid: {user_id}', key='SET')
            else:
                db.update_one(table=scfg.TABLE_NAME, sets=f'group_slug = \'{group_slug}\'',
                              condition=f'user_id = {user_id}')
                Debug(f'ReSet user group: {group_slug} uid: {user_id}', key='RST')
            self._send_message(user_id, cfg.SET_GROUP_TEXT.format(group_slug), 1)
            self._clear_wait_lists(user_id)  # Убираем из списка ожидания
            self._show_schedule_keyboard(user_id=user_id)  # Показываем клавиатуры выбора
            del db
        else:
            self._send_message(user_id, cfg.INVALID_GROUP_TEXT)
            Debug(f'Invalid group format {group_slug} uid: {user_id}', key='INV')

    def _clear_wait_lists(self, user_id: int) -> None:
        """
        Убирает пользователя из списка ожидания

        :param user_id:
        :return:
        """
        self.users_to_set_group.discard(str(user_id))
        self.users_to_set_teacher.discard(str(user_id))
        for i in range(len(self.users_to_get_teacher)):
            if self.users_to_get_teacher[i][0] == user_id:
                del self.users_to_get_teacher[i]
                break

    def _validate_teacher_name(self, teacher: str) -> bool:
        """
        Проверяет наличие преподавателя в файлах расписания

        :param teacher:
        :return:
        """
        for i in range(2, len(self.schedule_data), 4):
            for j in range(2, len(self.schedule_data[i])):
                tmp = self.schedule_data[i][j].split('\n')  # для сдвоенных пар
                if len(tmp) > 0:
                    # Исправляем отсутствие точки у некоторых преподавателей
                    t1 = tmp[0] if tmp[0][-1] == '.' else tmp[0] + '.'
                    t2 = tmp[-1] if tmp[-1][-1] == '.' else tmp[-1] + '.'
                    if t1 == teacher or t2 == teacher:
                        Debug(f'Find teacher {t1} or {t2}', key='FND')
                        return True
        return False

    def _validate_group_slug(self, group_slug: str) -> bool:
        """
        Проверка на валидность номера группы по маске и списку групп

        :param group_slug:
        :return:
        """
        group_slug = group_slug.upper()
        if re.match(scfg.GROUP_PATTERN, group_slug):
            if self._get_group_column(group_slug):
                return True
        return False

    def _reformat_corona_data(self, region: str, data: tuple) -> str:
        """
        Реформат данных коронавируса на 1 день

        :param region:
        :param data:
        :return:
        """
        status, value, plus = data
        return cfg.CORONA_STATS_PATTERN.format(status, region, value[0], plus[0], value[1], plus[1],
                                               value[2], plus[2], value[3], plus[3])

    def _reformat_subject_name(self, name: str or None, week_number: int, ignore_weeks: bool = False) -> str | None:
        """
        Реформат названия предмета с проверкой его присутствия на определённой неделе

        :param name:
        :param week_number:
        :param ignore_weeks:
        :return:
        """

        custom_week_pattern = r'кр. ([\d\,]+) н. ([^\\]+)'  # Кроме каких-то недель
        custom_week_range_pattern = r'(\d+\-\d+) н. ([^\\]+)'  # Диапазон
        custom_week_is_set_pattern = r'([\d\,]+) н. ([^\\]+)'  # Включая эти недели
        custom_week_dirt_pattern = r'…'  # Заглушки в расписании

        if name and name != 'None':  # Пара есть?
            data = name.split('\n')
            # Цикл, для сдвоенных пар
            for i in range(len(data)):
                if not ignore_weeks:
                    kr = re.search(custom_week_pattern, data[i])  # Проверяем, есть ли паттерн КР
                    if kr:
                        if str(week_number) in kr.group(1).split(','):  # Если неделя в списке исключённых удаляем
                            data[i] = cfg.WINDOW_SIGNATURE
                        else:
                            data[i] = kr.group(2)
                    else:
                        range_week = re.search(custom_week_range_pattern, data[i])
                        if range_week:
                            tmp = range_week.group(1).split('-')
                            from_week = int(tmp[0])
                            to_week = int(tmp[1])
                            if from_week <= week_number <= to_week:
                                data[i] = range_week.group(2)
                            else:
                                data[i] = cfg.WINDOW_SIGNATURE
                        else:
                            is_set = re.search(custom_week_is_set_pattern, data[i])
                            if is_set:
                                if str(week_number) in is_set.group(1).split(','):
                                    data[i] = is_set.group(2)
                                else:
                                    data[i] = cfg.WINDOW_SIGNATURE
                            else:
                                dirt = re.search(custom_week_dirt_pattern, data[i])
                                if dirt:
                                    data[i] = cfg.WINDOW_SIGNATURE
            return cfg.SPLIT_PAIR_SEPARATOR.join(data) if data else cfg.WINDOW_SIGNATURE
        return cfg.WINDOW_SIGNATURE

    def _reformat_double_pair(self, data: any) -> str:
        """
        Двойные и пустые пары в читабельный формат

        :param data:
        :return:
        """
        if data:
            if data == 'None':
                return ''
            return cfg.SPLIT_PAIR_SEPARATOR.join(data.split('\n'))
        return cfg.WINDOW_SIGNATURE

    def _reformat_day_schedule(self, data: list, date: datetime.datetime = datetime.datetime.now(),
                               week_format: bool = False, with_header: bool = True,
                               teacher_header: str | None = None) -> str:
        """
        Форматирует один день из списка в строку для дальнейшего вывода

        :param data:
        :param date:
        :param week_format:
        :param with_header:
        :param teacher_header:
        :return:
        """

        result = ''
        if with_header:
            if teacher_header:
                result += cfg.ONE_DAY_TEACHER_HEADER_PATTERN.format(
                    teacher_header, self._get_string_date(date, with_week_day=week_format))  # Дата
            else:
                result += cfg.ONE_DAY_HEADER_PATTERN.format(
                    self._get_string_date(date, with_week_day=week_format))  # Дата
        for i in range(len(data)):
            if len(data[i]) > 1:
                if data[i][0][:len(cfg.WINDOW_SIGNATURE)] != cfg.WINDOW_SIGNATURE:
                    result += cfg.ONE_PAIR_PATTERN.format(
                        i + 1,
                        str(data[i][0]),
                        str(data[i][1]) if data[i][1] != cfg.WINDOW_SIGNATURE and \
                                           data[i][1] != '' else cfg.VOID_SIGNATURE,
                        str(data[i][2]) if data[i][2] != cfg.WINDOW_SIGNATURE and \
                                           data[i][2] != '' else cfg.VOID_SIGNATURE,
                        str(data[i][3])) if data[i][3] != cfg.WINDOW_SIGNATURE and \
                                            data[i][3] != '' else cfg.VOID_SIGNATURE
                else:
                    result += cfg.ONE_PAIR_SHORT_PATTERN.format(i + 1, cfg.WINDOW_SIGNATURE)
            else:
                result += cfg.ONE_PAIR_SHORT_PATTERN.format(i + 1, cfg.WINDOW_SIGNATURE)
        return result

    def _reformat_teacher_name(self, teacher: str) -> str:
        """
        Форматирует имя преподавателя

        :param teacher:
        :return:
        """
        tmp = teacher.split(' ')
        if len(tmp) == 2:
            teacher = tmp[0].title() + ' ' + tmp[1].upper()
        elif len(tmp) == 1:
            teacher = tmp[0].title()
        return teacher

    def _send_message(self, user_id: int, text: str = '', keyboard: int = 0,
                      custom_keyboard: VkKeyboard = None) -> None:
        """
        Отправка сообщения

        :param user_id: ID пользователя
        :param text: Текс сообщения
        :param keyboard: 0 - Без клавиатуры(не менять), 1 - Стандартная клавиатура
        :return:
        """

        if keyboard == 1:
            keyboard = VkKeyboard(one_time=False)
            keyboard.add_button(cfg.BTN_SCHEDULE_TODAY, color=VkKeyboardColor.POSITIVE)
            keyboard.add_button(cfg.BTN_SCHEDULE_TOMORROW, color=VkKeyboardColor.NEGATIVE)
            keyboard.add_line()
            keyboard.add_button(cfg.BTN_SCHEDULE_WEEK, color=VkKeyboardColor.PRIMARY)
            keyboard.add_button(cfg.BTN_SCHEDULE_NEXT_WEEK, color=VkKeyboardColor.PRIMARY)
            keyboard.add_line()
            keyboard.add_button(cfg.BTN_WHAT_WEEK.title(), color=VkKeyboardColor.SECONDARY)
            keyboard.add_button(cfg.BTN_WHAT_GROUP.title(), color=VkKeyboardColor.SECONDARY)
            keyboard.add_button(cfg.BTN_HELP.title(), color=VkKeyboardColor.SECONDARY)
        if custom_keyboard:
            keyboard = custom_keyboard
        if keyboard:
            self.vk.messages.send(
                user_id=user_id,
                random_id=get_random_id(),
                keyboard=keyboard.get_keyboard(),
                message=text
            )
        else:
            self.vk.messages.send(
                user_id=user_id,
                random_id=get_random_id(),
                message=text
            )

    def _send_message_with_attachments(self, user_id: int, text: str = '', attachments: list = list) -> None:
        """
        Отправляет сообщение с вложениями

        :param user_id:
        :param text:
        :param attachments: Массив с вложениями
        :return:
        """
        self.vk.messages.send(
            user_id=user_id,
            attachment=','.join(attachments),
            random_id=get_random_id(),
            message=text
        )

    def _update_schedule_file(self) -> None:
        """
        Обновляет файл с расписанием

        :return:
        """
        page = requests.get(scfg.MIREA_SCHEDULE_URL)  # Получаем страницу
        soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
        result = soup.find(string="Институт информационных технологий").find_parent("div").find_parent("div").findAll(
            'a', {'class': 'uk-link-toggle'})
        course_pattern = r'([1-3]) курс'

        for x in result:
            course = x.find('div', 'uk-link-heading').text.lower().strip()
            course_number = re.match(course_pattern, course)
            if course_number:
                course_number = course_number.group(1)
                f = open(f'{scfg.DATA_DIR}{scfg.SCHEDULE_BASE_NAME}{course_number}.xlsx', "wb")
                link = x.get('href')
                resp = requests.get(link)  # запрос по ссылке
                f.write(resp.content)  # Записываем контент в файл
                f.close()

        self.last_schedule_file_update = time.time()
        Debug('Update schedule files', key='UPD')
        self._parse_schedule_file()

    def _parse_schedule_file(self) -> None:
        """
        Парсит полученные файлы расписание и записывает в списки

        :return:
        """
        self.schedule_data = []
        for c in range(3):
            book = openpyxl.load_workbook(
                f'{scfg.DATA_DIR}{scfg.SCHEDULE_BASE_NAME}{c + 1}.xlsx')  # открытие файла
            sheet = book.active  # активный лист
            num_cols = sheet.max_column  # количество столбцов
            last_group_cell = 0  # Сколько прошло ячеек от последней группы
            for i in range(6, num_cols):
                if last_group_cell >= 4:  # Если после группы прошло 4 ячейки, ждём следующей группы
                    last_group_cell = -1
                    continue
                column = []
                for j in range(2, 76):  # Перебираем
                    v = str(sheet.cell(column=i, row=j).value)
                    if j == 2 and re.match(scfg.GROUP_PATTERN,
                                           v):  # Если ячейка вторая, то проверяем что это номер группы
                        last_group_cell = 0  # Если это так, обнуляем счётчик
                    column.append(v)
                if last_group_cell != -1:  # Пока не дошли до следующей группы, не добавляем столбцы,
                    self.schedule_data.append(column)
                    last_group_cell += 1
        Debug('Parse schedule file', key='PRS')


def main():
    bot = VkBot()
    bot.start_listen()


if __name__ == '__main__':
    main()
